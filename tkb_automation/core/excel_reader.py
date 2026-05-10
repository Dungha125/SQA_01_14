"""Excel reader for loading test data from Excel files."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """Utility class to read test data from Excel (.xlsx) files."""

    def __init__(self, file_path: str | Path) -> None:
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

    def get_sheet_names(self) -> list[str]:
        """Return all sheet names in the workbook."""
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        return wb.sheetnames

    def read_sheet(self, sheet_name: str | None = None) -> list[dict[str, Any]]:
        """Read a sheet and return rows as list of dicts.

        Args:
            sheet_name: Name of the sheet. Uses the first sheet if None.

        Returns:
            List of row dicts with column names as keys.
        """
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in {self.file_path}")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        return self._sheet_to_dicts(ws)

    def read_row(self, row_idx: int, sheet_name: str | None = None) -> dict[str, Any] | None:
        """Read a specific row by index (1-based, header row is 1)."""
        rows = self.read_sheet(sheet_name)
        if 0 < row_idx <= len(rows):
            return rows[row_idx - 1]
        return None

    def _sheet_to_dicts(self, ws: Worksheet) -> list[dict[str, Any]]:
        """Convert a worksheet to a list of row dictionaries."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [self._normalize_header(str(cell) if cell is not None else "") for cell in rows[0]]
        result: list[dict[str, Any]] = []

        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            record: dict[str, Any] = {}
            for i, cell in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i}"
                if key:
                    record[key] = self._parse_cell(cell)
            result.append(record)

        return result

    def _normalize_header(self, header: str) -> str:
        """Normalize a column header to a snake_case key."""
        import re
        header = header.strip()
        header = re.sub(r"[^\w\s\-]", "", header)
        header = re.sub(r"\s+", "_", header)
        return header.lower()

    def _parse_cell(self, cell: Any) -> Any:
        """Parse a cell value, handling JSON strings, booleans, and numbers."""
        if cell is None:
            return None
        if isinstance(cell, str):
            s = cell.strip()
            if s.lower() == "true":
                return True
            if s.lower() == "false":
                return False
            if s.lower() == "none" or s.lower() == "null":
                return None
            if s.startswith("{") or s.startswith("["):
                try:
                    return json.loads(s)
                except json.JSONDecodeError:
                    return s
            return s
        if isinstance(cell, bool):
            return cell
        if isinstance(cell, (int, float)):
            return cell
        return str(cell)

    def get_tests_by_module(self, module_key: str) -> list[dict[str, Any]]:
        """Read a sheet named after the module key and return all rows."""
        try:
            return self.read_sheet(module_key)
        except ValueError:
            return []

    def filter_tests(
        self,
        sheet_name: str | None = None,
        test_type: str | None = None,
        auto_executable: str | None = None,
        priority: str | None = None,
    ) -> list[dict[str, Any]]:
        """Filter test cases from a sheet."""
        rows = self.read_sheet(sheet_name)
        if test_type:
            rows = [r for r in rows if str(r.get("test_type", "")).lower() == test_type.lower()]
        if auto_executable:
            rows = [
                r for r in rows
                if str(r.get("status", "")).lower() == auto_executable.lower()
            ]
        if priority:
            rows = [r for r in rows if str(r.get("priority", "")).lower() == priority.lower()]
        return rows
